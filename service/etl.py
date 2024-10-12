import pandas as pd
import os
import warnings

from helpers.excel import get_rows_rosco, there_are_columns_also_valid_and_not_valid, get_test_cells
from openpyxl import load_workbook
from helpers.utils import get_word_category

from models.definicion import Definicion
from models.base import Session
from sqlalchemy import delete

warnings.filterwarnings("ignore")

# Ruta a la carpeta que contiene los archivos Excel
ruta = 'xls'


# Iterar sobre cada archivo
def save_to_db():
    session = Session()
    session.execute(delete(Definicion))  # Ejecuta el comando de eliminación
    session.commit()  # Confirma los cambios
    resultados = []
    nro = 0
    for nombre in os.listdir(ruta):
        nro += 1
        if not (nombre.endswith('.xlsx') or nombre.endswith('.xls')) and not nombre.startswith('~$'):
            continue

        ruta_archivo = os.path.join(ruta, nombre)
        print(f'Chequeando archivo #{nro}')
        wb = load_workbook(ruta_archivo, data_only=True)

        for hoja in ['1', '2']:
            sheet_name = wb[hoja]
            celdas_testeos = get_test_cells(sheet_name)
            if not celdas_testeos:
                continue

            filas_rosco = get_rows_rosco(sheet_name)
            hay_tambien_valen_y_no_valen = there_are_columns_also_valid_and_not_valid(sheet_name, filas_rosco[0] - 1)

            for i in range(25):
                acepcion = sheet_name.cell(row=filas_rosco[i], column=3).value
                respuesta = sheet_name.cell(row=filas_rosco[i], column=4).value.capitalize() if sheet_name.cell(row=filas_rosco[i], column=4).value else None
                categoria_palabra = sheet_name.cell(row=filas_rosco[i], column=5).value.title() if sheet_name.cell(row=filas_rosco[i], column=5).value else None
                aciertos_testers = int(sheet_name.cell(row=celdas_testeos["filas"][i], column=celdas_testeos["columna"]).value)

                definicion_dict = {
                    'acepcion': acepcion,
                    'respuesta': respuesta,
                    'categoria_palabra': get_word_category(acepcion, categoria_palabra),
                    'tambien_valen': None,
                    'no_valen': None,
                    'aciertos_testers': aciertos_testers,
                }

                if hay_tambien_valen_y_no_valen:
                    definicion_dict['tambien_valen'] = sheet_name.cell(row=filas_rosco[i], column=6).value
                    definicion_dict['no_valen'] = sheet_name.cell(row=filas_rosco[i], column=7).value

                # resultados.append(definicion_dict)
                definicion = Definicion(**definicion_dict)
                session.add(definicion)  # Agregar la instancia a la sesión
                session.commit()
    # Guardar todos los cambios en la base de datos

    # Cerrar la sesión
    session.close()

    return resultados


def build_excel():
    session = Session()
    definiciones = session.query(Definicion).all()

    # Convertir las definiciones a un formato adecuado para pandas
    data = [
        {
            "Letra": definicion.respuesta[0],
            "Definición": definicion.acepcion,
            "Respuesta": definicion.respuesta,
            "Categoría de Palabra": definicion.categoria_palabra,
            "También Valen": definicion.tambien_valen,
            "NO Valen": definicion.no_valen,
            "Aciertos": definicion.aciertos_testers
        }
        for definicion in definiciones
    ]

    # Crear un DataFrame de pandas con los datos
    df = pd.DataFrame(data)

    # Guardar los datos en un archivo Excel
    output_path = "Archivo raiz con testeos.xlsx"
    df.to_excel(output_path, index=False)

    print(f"Definiciones guardadas en {output_path}")

    session.close()

    return {"status": 200, "msg": "ok"}